import openpyxl

# Открываем Excel файл
wb = openpyxl.load_workbook('Программа_подбора_ПК.xlsx', data_only=False)

# Получаем листы "Фреон" и "Таблица"
freon_sheet = wb['Фреон']
table_sheet = wb['Таблица']

# Диапазоны значений
d10_values = [round(16 + i * 0.5, 1) for i in range(int((45 - 16) / 0.5) + 1)]
d11_values = [20 + i for i in range(20, 71)]
d12_values = [round(25 + i * 0.5, 1) for i in range(int((52 - 25) / 0.5) + 1)]
d13_values = [20 + i for i in range(20, 71)]

# Создаем новый лист для всех результатов
result_sheet = wb.create_sheet('Результаты')

# Функция для копирования данных и добавления результатов
def add_result_to_sheet(d10, d11, d12, d13, row_offset):
    # Обновляем значения в ячейках на листе "Фреон"
    freon_sheet['D10'].value = d10
    freon_sheet['D11'].value = d11
    freon_sheet['D12'].value = d12
    freon_sheet['D13'].value = d13

    # Пересчитываем формулы в таблице
    wb.active = table_sheet
    wb.save('temp.xlsx')  # сохраняем изменения, чтобы Excel пересчитал формулы
    wb_recalculated = openpyxl.load_workbook('temp.xlsx', data_only=True)  # открываем с пересчитанными формулами
    recalculated_table_sheet = wb_recalculated['Таблица']

    # Копируем заголовки
    result_sheet.cell(row=row_offset, column=1, value=f'D10={d10}, D11={d11}, D12={d12}, D13={d13}')

    # Копируем данные из пересчитанной таблицы
    for row in recalculated_table_sheet.iter_rows(min_row=1, max_row=recalculated_table_sheet.max_row, min_col=1, max_col=recalculated_table_sheet.max_column):
        row_offset += 1
        for col_index, cell in enumerate(row, start=1):
            result_sheet.cell(row=row_offset, column=col_index, value=cell.value)

    # Добавляем пустую строку между результатами
    return row_offset + 1

# Начальная строка для результатов
current_row = 1

# Проходим по всем комбинациям значений
for d10 in d10_values:
    for d11 in d11_values:
        for d12 in d12_values:
            for d13 in d13_values:
                current_row = add_result_to_sheet(d10, d11, d12, d13, current_row)

# Сохраняем результирующий Excel файл
wb.save('Программа_подбора_ПК_результавфцвты.xlsx')

print('Файл с результатами сохранен!')